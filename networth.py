"""
networth.py

Net Worth Tracking Tool (Streamlit)

Core pillars:
1) Net Worth = Assets - Debts
2) Monthly Cash Flow = Income - Expenses (with Essential toggle)
3) Financial Runway = Emergency Fund / Essential Expenses

Complementary status label:
- Stable / Vulnerable / High Stress
Good debt rule:
- interest < 4% annually OR debt type == Mortgage

Exports:
- Excel (.xlsx): Summary + Assets + Debts + Cash Flow sheets (requires openpyxl)
- PDF (.pdf): Summary + Tables (Assets, Debts, Income, Expenses) (requires reportlab)

IMPORTANT:
Your error indicates your local Python environment likely doesn't have openpyxl installed.
This script handles missing dependencies gracefully and will show install instructions in the UI.
"""

from __future__ import annotations

import datetime as _dt
from io import BytesIO
from typing import Any, Dict, List, Tuple

import streamlit as st
from storage import signup_save, login_save

# -------------------------
# Optional dependencies
# -------------------------
OPENPYXL_AVAILABLE = True
REPORTLAB_AVAILABLE = True
OPENPYXL_ERR = ""
REPORTLAB_ERR = ""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception as e:  # ModuleNotFoundError or others
    OPENPYXL_AVAILABLE = False
    OPENPYXL_ERR = str(e)

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )
except Exception as e:
    REPORTLAB_AVAILABLE = False
    REPORTLAB_ERR = str(e)


# =========================
# Constants and styling
# =========================

CURRENCY_SYMBOL = {"EUR": "€", "IDR": "Rp", "CNY": "¥"}

PALETTE = {
    "cream": "#FAF2DE",
    "paper": "#FFF9EA",
    "teal": "#1F8A86",
    "mustard": "#E0B12E",
    "ink": "#1B1B1B",
}


def _inject_retro_css() -> None:
    css = f"""
    <style>
      :root {{
        --cream: {PALETTE['cream']};
        --paper: {PALETTE['paper']};
        --teal: {PALETTE['teal']};
        --mustard: {PALETTE['mustard']};
        --ink: {PALETTE['ink']};
      }}

      .stApp {{
        background: var(--paper);
        color: var(--ink);
      }}

      .stButton > button {{
        border-radius: 14px;
        border: 2px solid var(--ink);
        box-shadow: 2px 2px 0 var(--ink);
      }}

      .stButton > button[kind="primary"] {{
        background: var(--teal);
        color: white;
      }}

      [data-testid="stMetric"] {{
        background: var(--cream);
        border: 2px solid var(--ink);
        border-radius: 16px;
        padding: 12px 14px;
        box-shadow: 3px 3px 0 var(--ink);
      }}

      button[role="tab"] {{
        border-radius: 14px;
        border: 2px solid var(--ink);
        margin-right: 6px;
      }}

      [data-baseweb="input"] > div {{
        border-radius: 14px;
      }}
      [data-baseweb="textarea"] > div {{
        border-radius: 14px;
      }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)


def _card(title: str, body_html: str, bg_color_key: str = "cream") -> None:
    bg = PALETTE.get(bg_color_key, PALETTE["cream"])
    st.markdown(
        f"""
        <div style="background:{bg}; border:2px solid {PALETTE['ink']}; border-radius:18px;
                    padding:14px 16px; box-shadow:3px 3px 0 {PALETTE['ink']}; margin-bottom:10px;">
          <div style="font-weight:800; font-size:18px; margin-bottom:6px;">{title}</div>
          <div style="font-size:14px; line-height:1.4;">{body_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# =========================
# Helpers
# =========================

def _today_iso() -> str:
    return _dt.date.today().isoformat()


def _currency_step(cur: str) -> float:
    return 500_000.0 if cur == "IDR" else 100.0


def _fmt_money(value: float, cur: str) -> str:
    sym = CURRENCY_SYMBOL.get(cur, "")
    v = float(value or 0.0)
    if cur == "IDR":
        return f"{sym}{v:,.0f}".replace(",", ".")
    return f"{sym}{v:,.2f}"


def _safe_float(x: Any, default: float = 0.0) -> float:
    try:
        return float(x)
    except Exception:
        return float(default)


def _sum(items: List[Dict[str, Any]], key: str) -> float:
    return float(sum(_safe_float(it.get(key, 0.0), 0.0) for it in items))


# =========================
# Business logic
# =========================

def is_good_debt(debt: Dict[str, Any]) -> bool:
    d_type = str(debt.get("type", "")).strip().lower()
    if d_type == "mortgage":
        return True
    r = _safe_float(debt.get("interest", 0.0), 0.0)
    return r < 4.0


def compute_status(
    assets_total: float,
    debts_total: float,
    bad_debt_total: float,
    cashflow: float,
    emergency_fund: float,
    essential_monthly: float,
) -> Tuple[str, str]:
    A = float(assets_total or 0.0)
    D = float(debts_total or 0.0)
    BD = float(bad_debt_total or 0.0)
    CF = float(cashflow or 0.0)
    EF = float(emergency_fund or 0.0)
    ESS = float(essential_monthly or 0.0)

    if D > A:
        return "High Stress", "Debt is higher than assets. Prioritize reducing debt and stabilizing your balance sheet."
    if EF <= 0 and CF <= 0:
        return "High Stress", "No emergency fund and cash flow is not positive. Build liquidity first."

    runway = (EF / ESS) if ESS > 0 else 0.0
    stable_ok = runway >= 3.0 and CF > 0 and (D == 0.0 or (D <= A and BD == 0.0))
    if stable_ok:
        return "Stable", "Positive cash flow and a 3+ month buffer. Debt looks healthy."
    return "Vulnerable", "You are mostly ok, but your buffer, cash flow, or debt type needs attention."


# =========================
# Session state
# =========================

def _init_state() -> None:
    # Data
    st.session_state.setdefault("nw_assets_items", [])
    st.session_state.setdefault("nw_debts_items", [])
    st.session_state.setdefault("nw_cf_month", _dt.date.today().strftime("%Y-%m"))
    st.session_state.setdefault("nw_income_items", [])
    st.session_state.setdefault("nw_expense_items", [])
    st.session_state.setdefault("nw_emergency_fund", 0.0)  # SAME key used in dashboard + runway tab
    st.session_state.setdefault("nw_note", "")

    # Form fields (safe to clear in callbacks)
    st.session_state.setdefault("nw_asset_cat", "Cash")
    st.session_state.setdefault("nw_asset_name", "")
    st.session_state.setdefault("nw_asset_value", 0.0)

    st.session_state.setdefault("nw_debt_type", "Mortgage")
    st.session_state.setdefault("nw_debt_name", "")
    st.session_state.setdefault("nw_debt_balance", 0.0)
    st.session_state.setdefault("nw_debt_interest", 0.0)

    st.session_state.setdefault("nw_income_name", "")
    st.session_state.setdefault("nw_income_value", 0.0)

    st.session_state.setdefault("nw_expense_name", "")
    st.session_state.setdefault("nw_expense_value", 0.0)
    st.session_state.setdefault("nw_expense_essential", True)


def _reset_all() -> None:
    st.session_state["nw_assets_items"] = []
    st.session_state["nw_debts_items"] = []
    st.session_state["nw_income_items"] = []
    st.session_state["nw_expense_items"] = []
    st.session_state["nw_emergency_fund"] = 0.0
    st.session_state["nw_note"] = ""
    st.session_state["nw_cf_month"] = _dt.date.today().strftime("%Y-%m")

    st.session_state["nw_asset_cat"] = "Cash"
    st.session_state["nw_asset_name"] = ""
    st.session_state["nw_asset_value"] = 0.0

    st.session_state["nw_debt_type"] = "Mortgage"
    st.session_state["nw_debt_name"] = ""
    st.session_state["nw_debt_balance"] = 0.0
    st.session_state["nw_debt_interest"] = 0.0

    st.session_state["nw_income_name"] = ""
    st.session_state["nw_income_value"] = 0.0

    st.session_state["nw_expense_name"] = ""
    st.session_state["nw_expense_value"] = 0.0
    st.session_state["nw_expense_essential"] = True


# =========================
# Callbacks (safe state edits)
# =========================

def _cb_add_asset() -> None:
    name = (st.session_state.get("nw_asset_name") or "").strip()
    if not name:
        st.session_state["nw_toast_error"] = "Asset name is required."
        return
    st.session_state["nw_assets_items"].append(
        {
            "category": st.session_state.get("nw_asset_cat", "Other"),
            "name": name,
            "value": float(_safe_float(st.session_state.get("nw_asset_value", 0.0), 0.0)),
        }
    )
    st.session_state["nw_asset_name"] = ""
    st.session_state["nw_asset_value"] = 0.0


def _cb_add_debt() -> None:
    name = (st.session_state.get("nw_debt_name") or "").strip()
    if not name:
        st.session_state["nw_toast_error"] = "Debt name is required."
        return
    st.session_state["nw_debts_items"].append(
        {
            "type": st.session_state.get("nw_debt_type", "Other"),
            "name": name,
            "balance": float(_safe_float(st.session_state.get("nw_debt_balance", 0.0), 0.0)),
            "interest": float(_safe_float(st.session_state.get("nw_debt_interest", 0.0), 0.0)),
        }
    )
    st.session_state["nw_debt_name"] = ""
    st.session_state["nw_debt_balance"] = 0.0
    st.session_state["nw_debt_interest"] = 0.0


def _cb_add_income() -> None:
    name = (st.session_state.get("nw_income_name") or "").strip()
    if not name:
        st.session_state["nw_toast_error"] = "Income name is required."
        return
    st.session_state["nw_income_items"].append(
        {"name": name, "value": float(_safe_float(st.session_state.get("nw_income_value", 0.0), 0.0))}
    )
    st.session_state["nw_income_name"] = ""
    st.session_state["nw_income_value"] = 0.0


def _cb_add_expense() -> None:
    name = (st.session_state.get("nw_expense_name") or "").strip()
    if not name:
        st.session_state["nw_toast_error"] = "Expense name is required."
        return
    st.session_state["nw_expense_items"].append(
        {
            "name": name,
            "value": float(_safe_float(st.session_state.get("nw_expense_value", 0.0), 0.0)),
            "essential": bool(st.session_state.get("nw_expense_essential", True)),
        }
    )
    st.session_state["nw_expense_name"] = ""
    st.session_state["nw_expense_value"] = 0.0
    st.session_state["nw_expense_essential"] = True


def _flush_toast() -> None:
    msg = st.session_state.pop("nw_toast_error", None)
    if msg:
        st.warning(msg)


# =========================
# Export builders
# =========================

def _make_payload(cur: str) -> Dict[str, Any]:
    assets_items: List[Dict[str, Any]] = st.session_state["nw_assets_items"]
    debts_items: List[Dict[str, Any]] = st.session_state["nw_debts_items"]
    income_items: List[Dict[str, Any]] = st.session_state["nw_income_items"]
    expense_items: List[Dict[str, Any]] = st.session_state["nw_expense_items"]

    assets_total = _sum(assets_items, "value")
    debts_total = _sum(debts_items, "balance")
    net_worth = assets_total - debts_total

    income_total = _sum(income_items, "value")
    expense_total = _sum(expense_items, "value")
    cashflow = income_total - expense_total

    essential_total = _sum([e for e in expense_items if e.get("essential")], "value")
    emergency_fund = _safe_float(st.session_state["nw_emergency_fund"], 0.0)
    runway_months = (emergency_fund / essential_total) if essential_total > 0 else 0.0

    bad_debt_total = sum(
        _safe_float(d.get("balance", 0.0), 0.0) for d in debts_items if not is_good_debt(d)
    )

    status, status_msg = compute_status(
        assets_total=assets_total,
        debts_total=debts_total,
        bad_debt_total=bad_debt_total,
        cashflow=cashflow,
        emergency_fund=emergency_fund,
        essential_monthly=essential_total,
    )

    return {
        "date": _today_iso(),
        "month": st.session_state["nw_cf_month"],
        "currency": cur,
        "assets_items": assets_items,
        "debts_items": debts_items,
        "income_items": income_items,
        "expense_items": expense_items,
        "emergency_fund": float(emergency_fund),
        "assets_total": float(assets_total),
        "debts_total": float(debts_total),
        "net_worth": float(net_worth),
        "income_total": float(income_total),
        "expense_total": float(expense_total),
        "cashflow": float(cashflow),
        "essential_expense_total": float(essential_total),
        "runway_months": float(runway_months),
        "bad_debt_total": float(bad_debt_total),
        "status": status,
        "status_msg": status_msg,
        "note": st.session_state.get("nw_note", ""),
    }


# -------- Excel --------
def _autosize_sheet(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 48)


def build_excel_bytes(payload: Dict[str, Any]) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError(f"openpyxl is not available: {OPENPYXL_ERR}")

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    center = Alignment(vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Field", "Value"])
    ws["A1"].font = header_font
    ws["B1"].font = header_font

    rows = [
        ("Date", payload.get("date", "")),
        ("Month", payload.get("month", "")),
        ("Currency", payload.get("currency", "")),
        ("Assets total", payload.get("assets_total", 0.0)),
        ("Debts total", payload.get("debts_total", 0.0)),
        ("Net worth", payload.get("net_worth", 0.0)),
        ("Income total", payload.get("income_total", 0.0)),
        ("Expense total", payload.get("expense_total", 0.0)),
        ("Cashflow", payload.get("cashflow", 0.0)),
        ("Emergency fund", payload.get("emergency_fund", 0.0)),
        ("Essential expense total", payload.get("essential_expense_total", 0.0)),
        ("Runway months", payload.get("runway_months", 0.0)),
        ("Status", payload.get("status", "")),
        ("Status message", payload.get("status_msg", "")),
        ("Note", payload.get("note", "")),
    ]
    for k, v in rows:
        ws.append([k, v])

    for r in range(2, ws.max_row + 1):
        ws[f"A{r}"].alignment = center
        ws[f"B{r}"].alignment = right
    _autosize_sheet(ws)

    # Assets
    ws = wb.create_sheet("Assets")
    ws.append(["Category", "Name", "Value"])
    for cell in ws[1]:
        cell.font = header_font
    for a in payload.get("assets_items", []):
        ws.append([a.get("category", ""), a.get("name", ""), a.get("value", 0.0)])
    _autosize_sheet(ws)

    # Debts
    ws = wb.create_sheet("Debts")
    ws.append(["Type", "Name", "Balance", "Interest (%)", "Good debt?"])
    for cell in ws[1]:
        cell.font = header_font
    for d in payload.get("debts_items", []):
        good = "Yes" if is_good_debt(d) else "No"
        ws.append([d.get("type", ""), d.get("name", ""), d.get("balance", 0.0), d.get("interest", 0.0), good])
    _autosize_sheet(ws)

    # Cash Flow
    ws = wb.create_sheet("Cash Flow")
    ws.append(["Section", "Name", "Value", "Essential?"])
    for cell in ws[1]:
        cell.font = header_font
    for inc in payload.get("income_items", []):
        ws.append(["Income", inc.get("name", ""), inc.get("value", 0.0), ""])
    for exp in payload.get("expense_items", []):
        ws.append(["Expense", exp.get("name", ""), exp.get("value", 0.0), "Yes" if exp.get("essential") else "No"])
    _autosize_sheet(ws)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -------- PDF --------
def _pdf_table(title: str, rows: List[List[Any]]) -> List[Any]:
    styles = getSampleStyleSheet()
    elems: List[Any] = []
    elems.append(Paragraph(f"<b>{title}</b>", styles["Heading3"]))
    elems.append(Spacer(1, 6))

    if not rows:
        elems.append(Paragraph("No data.", styles["Normal"]))
        elems.append(Spacer(1, 12))
        return elems

    data = [[str(x) if x is not None else "" for x in r] for r in rows]
    t = Table(data, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightyellow]),
            ]
        )
    )
    elems.append(t)
    elems.append(Spacer(1, 14))
    return elems


def build_pdf_bytes(payload: Dict[str, Any]) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError(f"reportlab is not available: {REPORTLAB_ERR}")

    styles = getSampleStyleSheet()
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, title="Net Worth Snapshot")

    cur = payload.get("currency", "EUR")

    def money(v: Any) -> str:
        return _fmt_money(_safe_float(v, 0.0), cur)

    elems: List[Any] = []
    elems.append(Paragraph("<b>Net Worth Snapshot</b>", styles["Title"]))
    elems.append(
        Paragraph(
            f"Date: {payload.get('date','')} | Month: {payload.get('month','')} | Currency: {cur}",
            styles["Normal"],
        )
    )
    elems.append(Spacer(1, 12))

    summary = [
        ["Metric", "Value"],
        ["Assets total", money(payload.get("assets_total", 0.0))],
        ["Debts total", money(payload.get("debts_total", 0.0))],
        ["Net worth", money(payload.get("net_worth", 0.0))],
        ["Income total", money(payload.get("income_total", 0.0))],
        ["Expense total", money(payload.get("expense_total", 0.0))],
        ["Cashflow", money(payload.get("cashflow", 0.0))],
        ["Emergency fund", money(payload.get("emergency_fund", 0.0))],
        ["Essential expense total", money(payload.get("essential_expense_total", 0.0)) + " / mo"],
        ["Runway months", f"{_safe_float(payload.get('runway_months',0.0),0.0):.1f}"],
        ["Status", str(payload.get("status", ""))],
    ]
    elems += _pdf_table("Summary", summary)

    status_msg = str(payload.get("status_msg", "")).strip()
    if status_msg:
        elems.append(Paragraph(f"<b>Status note:</b> {status_msg}", styles["Normal"]))
        elems.append(Spacer(1, 10))

    note = str(payload.get("note", "")).strip()
    if note:
        elems.append(Paragraph(f"<b>User note:</b> {note}", styles["Normal"]))
        elems.append(Spacer(1, 12))

    assets_rows = [["Category", "Name", "Value"]]
    for a in payload.get("assets_items", []):
        assets_rows.append([a.get("category", ""), a.get("name", ""), money(a.get("value", 0.0))])

    debts_rows = [["Type", "Name", "Balance", "Interest (%)", "Good debt?"]]
    for d in payload.get("debts_items", []):
        debts_rows.append(
            [
                d.get("type", ""),
                d.get("name", ""),
                money(d.get("balance", 0.0)),
                f"{_safe_float(d.get('interest',0.0),0.0):.1f}",
                "Yes" if is_good_debt(d) else "No",
            ]
        )

    income_rows = [["Name", "Value"]]
    for i in payload.get("income_items", []):
        income_rows.append([i.get("name", ""), money(i.get("value", 0.0))])

    expense_rows = [["Name", "Value", "Essential?"]]
    for e in payload.get("expense_items", []):
        expense_rows.append([e.get("name", ""), money(e.get("value", 0.0)), "Yes" if e.get("essential") else "No"])

    elems.append(PageBreak())
    elems += _pdf_table("Assets", assets_rows)
    elems += _pdf_table("Debts", debts_rows)
    elems += _pdf_table("Income", income_rows)
    elems += _pdf_table("Expenses", expense_rows)

    doc.build(elems)
    return bio.getvalue()


# =========================
# UI sections
# =========================

def _dashboard_tab(cur: str, on_back) -> None:
    assets_items = st.session_state["nw_assets_items"]
    debts_items = st.session_state["nw_debts_items"]
    income_items = st.session_state["nw_income_items"]
    expense_items = st.session_state["nw_expense_items"]

    assets_total = _sum(assets_items, "value")
    debts_total = _sum(debts_items, "balance")
    net_worth = assets_total - debts_total

    income_total = _sum(income_items, "value")
    expense_total = _sum(expense_items, "value")
    cashflow = income_total - expense_total

    essential_total = _sum([e for e in expense_items if e.get("essential")], "value")
    emergency_fund = _safe_float(st.session_state["nw_emergency_fund"], 0.0)
    runway_months = (emergency_fund / essential_total) if essential_total > 0 else 0.0

    bad_debt_total = sum(_safe_float(d.get("balance", 0.0), 0.0) for d in debts_items if not is_good_debt(d))

    status, status_msg = compute_status(
        assets_total=assets_total,
        debts_total=debts_total,
        bad_debt_total=bad_debt_total,
        cashflow=cashflow,
        emergency_fund=emergency_fund,
        essential_monthly=essential_total,
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Net worth", _fmt_money(net_worth, cur))
    with c2:
        st.metric("Cash flow", _fmt_money(cashflow, cur))
    with c3:
        st.metric("Runway", f"{runway_months:.1f} mo" if essential_total > 0 else "Set essentials")

    _card(
        "Pillar 1: Net worth",
        f"Assets: <b>{_fmt_money(assets_total, cur)}</b><br>Debts: <b>{_fmt_money(debts_total, cur)}</b>",
        "cream",
    )
    _card(
        "Pillar 2: Cash flow (this month)",
        f"Income: <b>{_fmt_money(income_total, cur)}</b><br>Expenses: <b>{_fmt_money(expense_total, cur)}</b>",
        "cream",
    )

    if essential_total <= 0:
        runway_html = f"Emergency fund: <b>{_fmt_money(emergency_fund, cur)}</b><br>Essential expenses: <b>not set</b>"
    else:
        runway_html = (
            f"Emergency fund: <b>{_fmt_money(emergency_fund, cur)}</b><br>"
            f"Essential expenses: <b>{_fmt_money(essential_total, cur)}/mo</b><br>"
            f"Runway: <b>{runway_months:.1f} months</b>"
        )
    _card("Pillar 3: Financial runway", runway_html, "cream")

    status_color = "teal" if status == "Stable" else "mustard"
    _card(f"Status: {status}", status_msg, status_color)

    st.divider()
    b1, b2 = st.columns(2)
    with b1:
        st.button("Back to home", on_click=on_back, key="nw_btn_back_home_dashboard")
    with b2:
        st.button("Reset all", on_click=_reset_all, key="nw_btn_reset_all_dashboard")


def _networth_tab(cur: str) -> None:
    step_val = _currency_step(cur)
    assets_items: List[Dict[str, Any]] = st.session_state["nw_assets_items"]
    debts_items: List[Dict[str, Any]] = st.session_state["nw_debts_items"]

    _flush_toast()

    st.subheader("Assets")
    with st.expander("Add asset", expanded=True):
        col1, col2 = st.columns([1, 2])
        with col1:
            st.selectbox("Category", ["Cash", "Investments", "Property", "Other"], key="nw_asset_cat")
        with col2:
            st.text_input("Name", placeholder="Savings account, ETF portfolio, Home value", key="nw_asset_name")
        st.number_input(f"Value ({cur})", min_value=0.0, step=step_val, key="nw_asset_value")
        st.button("Add asset", type="primary", key="nw_btn_add_asset", on_click=_cb_add_asset)

    if assets_items:
        for i, a in enumerate(list(assets_items)):
            row = st.columns([2, 3, 2, 1])
            row[0].write(a.get("category", ""))
            row[1].write(a.get("name", ""))
            row[2].write(_fmt_money(_safe_float(a.get("value", 0.0)), cur))
            if row[3].button("✕", key=f"nw_btn_del_asset_{i}"):
                assets_items.pop(i)
                st.rerun()
    else:
        st.info("No assets yet. Add your first asset above.")

    st.divider()
    st.subheader("Debts")
    with st.expander("Add debt", expanded=True):
        col1, col2 = st.columns([1, 2])
        with col1:
            st.selectbox("Type", ["Mortgage", "Loan", "Credit card", "Other"], key="nw_debt_type")
        with col2:
            st.text_input("Name", placeholder="Mortgage ABN, Car loan, Credit card", key="nw_debt_name")

        col3, col4 = st.columns(2)
        with col3:
            st.number_input(f"Balance ({cur})", min_value=0.0, step=step_val, key="nw_debt_balance")
        with col4:
            st.number_input("Interest rate (% per year)", min_value=0.0, max_value=100.0, step=0.1, key="nw_debt_interest")

        st.button("Add debt", type="primary", key="nw_btn_add_debt", on_click=_cb_add_debt)

    if debts_items:
        for i, d in enumerate(list(debts_items)):
            good = is_good_debt(d)
            tag = "Good" if good else "Bad"
            tag_color = PALETTE["teal"] if good else PALETTE["mustard"]

            st.markdown(
                f"""
                <div style="background:{PALETTE['cream']}; border:2px solid {PALETTE['ink']}; border-radius:16px;
                            padding:10px 12px; box-shadow:2px 2px 0 {PALETTE['ink']}; margin-bottom:8px;">
                  <div style="display:flex; justify-content:space-between; align-items:center; gap:10px;">
                    <div>
                      <div style="font-weight:800;">{d.get('name','')}</div>
                      <div style="font-size:13px;">
                        {d.get('type','')} • {_safe_float(d.get('interest',0.0),0.0):.1f}%
                      </div>
                    </div>
                    <div style="text-align:right;">
                      <div style="font-weight:800;">{_fmt_money(_safe_float(d.get('balance',0.0),0.0), cur)}</div>
                      <div style="display:inline-block; padding:2px 10px; border-radius:999px; background:{tag_color};
                                  color:white; font-size:12px; font-weight:700;">{tag}</div>
                    </div>
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.button("Remove", key=f"nw_btn_del_debt_{i}"):
                debts_items.pop(i)
                st.rerun()
    else:
        st.info("No debts yet. Add your first debt above.")

    st.divider()
    assets_total = _sum(assets_items, "value")
    debts_total = _sum(debts_items, "balance")
    net_worth = assets_total - debts_total

    st.subheader("Summary")
    st.write(
        f"**Assets:** {_fmt_money(assets_total, cur)}  \n"
        f"**Debts:** {_fmt_money(debts_total, cur)}  \n"
        f"**Net worth:** {_fmt_money(net_worth, cur)}"
    )


def _cashflow_tab(cur: str) -> None:
    step_val = _currency_step(cur)
    income_items: List[Dict[str, Any]] = st.session_state["nw_income_items"]
    expense_items: List[Dict[str, Any]] = st.session_state["nw_expense_items"]

    _flush_toast()

    st.subheader("Monthly cash flow")
    st.caption("Track income and expenses for the selected month. Essentials are used for runway.")

    month = st.text_input("Month (YYYY-MM)", value=st.session_state["nw_cf_month"], key="nw_cf_month_input")
    st.session_state["nw_cf_month"] = month.strip() if month else st.session_state["nw_cf_month"]

    st.markdown("#### Income")
    i1, i2, i3 = st.columns([3, 2, 1])
    with i1:
        st.text_input("Name", placeholder="Salary, Side hustle", key="nw_income_name")
    with i2:
        st.number_input(f"Amount ({cur})", min_value=0.0, step=step_val, key="nw_income_value")
    with i3:
        st.button("Add", type="primary", key="nw_btn_add_income", on_click=_cb_add_income)

    if income_items:
        for i, it in enumerate(list(income_items)):
            row = st.columns([4, 2, 1])
            row[0].write(it.get("name", ""))
            row[1].write(_fmt_money(_safe_float(it.get("value", 0.0)), cur))
            if row[2].button("✕", key=f"nw_btn_del_income_{i}"):
                income_items.pop(i)
                st.rerun()
    else:
        st.info("No income items yet.")

    st.markdown("#### Expenses")
    e1, e2, e3, e4 = st.columns([3, 2, 1, 1])
    with e1:
        st.text_input("Name ", placeholder="Rent, Groceries", key="nw_expense_name")
    with e2:
        st.number_input(f"Amount  ({cur})", min_value=0.0, step=step_val, key="nw_expense_value")
    with e3:
        st.checkbox("Essential", value=True, key="nw_expense_essential")
    with e4:
        st.button("Add ", type="primary", key="nw_btn_add_expense", on_click=_cb_add_expense)

    if expense_items:
        for i, it in enumerate(list(expense_items)):
            row = st.columns([4, 2, 2, 1])
            row[0].write(it.get("name", ""))
            row[1].write(_fmt_money(_safe_float(it.get("value", 0.0)), cur))
            row[2].write("Essential" if it.get("essential") else "Discretionary")
            if row[3].button("✕", key=f"nw_btn_del_expense_{i}"):
                expense_items.pop(i)
                st.rerun()
    else:
        st.info("No expense items yet.")

    st.divider()
    income_total = _sum(income_items, "value")
    expense_total = _sum(expense_items, "value")
    cashflow = income_total - expense_total
    essential_total = _sum([e for e in expense_items if e.get("essential")], "value")

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Income", _fmt_money(income_total, cur))
    with m2:
        st.metric("Expenses", _fmt_money(expense_total, cur))
    with m3:
        st.metric("Cash flow", _fmt_money(cashflow, cur))

    st.caption(f"Essential expenses used for runway: {_fmt_money(essential_total, cur)}/mo")


def _runway_tab(cur: str) -> None:
    step_val = _currency_step(cur)
    expense_items: List[Dict[str, Any]] = st.session_state["nw_expense_items"]
    essential_total = _sum([e for e in expense_items if e.get("essential")], "value")

    st.subheader("Emergency fund and runway")
    st.caption("Emergency fund is separate and should be liquid cash.")

    st.number_input(
        f"Emergency fund (liquid cash) ({cur})",
        min_value=0.0,
        step=step_val,
        key="nw_emergency_fund",
    )
    ef = _safe_float(st.session_state["nw_emergency_fund"], 0.0)

    if essential_total <= 0:
        st.warning("Set at least one expense as Essential in Cash Flow, otherwise runway cannot be calculated.")
        st.metric("Runway", "Set essentials")
        runway_months = 0.0
    else:
        runway_months = ef / essential_total
        st.metric("Essential expenses", f"{_fmt_money(essential_total, cur)}/mo")
        st.metric("Runway", f"{runway_months:.1f} months")

    st.divider()
    st.subheader("Benchmarks")
    b1, b2, b3 = st.columns(3)
    with b1:
        st.write("3 months")
        st.write("✅" if runway_months >= 3 else "⏳")
    with b2:
        st.write("6 months")
        st.write("✅" if runway_months >= 6 else "⏳")
    with b3:
        st.write("12 months")
        st.write("✅" if runway_months >= 12 else "⏳")


def _save_export_tab(cur: str, on_back) -> None:
    st.subheader("Save snapshot and export")
    st.caption("Exports include summary + tables. Excel is for analysis, PDF is for sharing.")

    note = st.text_input("Note (optional)", value=st.session_state["nw_note"], key="nw_note_input")
    st.session_state["nw_note"] = note

    payload = _make_payload(cur)

    # Dependency notice
    if not OPENPYXL_AVAILABLE or not REPORTLAB_AVAILABLE:
        missing = []
        if not OPENPYXL_AVAILABLE:
            missing.append(f"openpyxl ({OPENPYXL_ERR})")
        if not REPORTLAB_AVAILABLE:
            missing.append(f"reportlab ({REPORTLAB_ERR})")
        st.warning(
            "Some export options are disabled because dependencies are missing:\n\n"
            + "\n".join([f"- {m}" for m in missing])
            + "\n\nInstall them with:\n"
            "pip install openpyxl reportlab"
        )

    st.markdown("#### Export")

    c1, c2 = st.columns(2)
    with c1:
        if OPENPYXL_AVAILABLE:
            xlsx_bytes = build_excel_bytes(payload)
            st.download_button(
                "Download Excel (.xlsx)",
                data=xlsx_bytes,
                file_name=f"networth_snapshot_{payload['date']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="nw_dl_excel",
            )
        else:
            st.button("Download Excel (.xlsx)", key="nw_dl_excel_disabled", disabled=True)

    with c2:
        if REPORTLAB_AVAILABLE:
            pdf_bytes = build_pdf_bytes(payload)
            st.download_button(
                "Download PDF (.pdf)",
                data=pdf_bytes,
                file_name=f"networth_snapshot_{payload['date']}.pdf",
                mime="application/pdf",
                key="nw_dl_pdf",
            )
        else:
            st.button("Download PDF (.pdf)", key="nw_dl_pdf_disabled", disabled=True)

    st.divider()
    tab_su, tab_li = st.tabs(["Sign up and Save", "Log in and Save"])

    with tab_su:
        email = st.text_input("Email", key="nw_signup_email")
        pw = st.text_input("Password", type="password", key="nw_signup_pw")
        if st.button("Sign up and Save", type="primary", key="nw_btn_signup_save"):
            ok, msg = signup_save(email, pw, payload, record_type="networth")
            st.success(msg) if ok else st.error(msg)

    with tab_li:
        email2 = st.text_input("Email ", key="nw_login_email")
        pw2 = st.text_input("Password ", type="password", key="nw_login_pw")
        if st.button("Log in and Save", type="primary", key="nw_btn_login_save"):
            ok, msg = login_save(email2, pw2, payload, record_type="networth")
            st.success(msg) if ok else st.error(msg)

    st.divider()
    _card(
        f"Status: {payload.get('status','')}",
        str(payload.get("status_msg", "")),
        "teal" if payload.get("status") == "Stable" else "mustard",
    )

    st.divider()
    c3, c4 = st.columns(2)
    with c3:
        st.button("Back to home", on_click=on_back, key="nw_btn_back_home_save_export")
    with c4:
        if st.button("Reset note", key="nw_btn_reset_note"):
            st.session_state["nw_note"] = ""
            st.rerun()


# =========================
# Public entry point
# =========================

def render_networth(on_back) -> None:
    _init_state()
    _inject_retro_css()

    cur = st.session_state.get("currency", "EUR")

    st.title("Net Worth Tracker")
    st.caption("Tracks net worth, monthly cash flow, and financial runway. Status is complementary.")

    tabs = st.tabs(["Dashboard", "Net Worth", "Cash Flow", "Runway", "Save & Export"])

    with tabs[0]:
        _dashboard_tab(cur, on_back)
    with tabs[1]:
        _networth_tab(cur)
    with tabs[2]:
        _cashflow_tab(cur)
    with tabs[3]:
        _runway_tab(cur)
    with tabs[4]:
        _save_export_tab(cur, on_back)